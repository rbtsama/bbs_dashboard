import os
import pandas as pd
from datetime import datetime, timedelta
from tqdm import tqdm
from colorama import init, Fore
import glob
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

# 初始化colorama
init()

# 设置数据路径
DATA_DIR = Path(__file__).parent.parent / 'data'
RAW_DIR = DATA_DIR / 'raw'
PROCESSED_DIR = DATA_DIR / 'processed'

def round_time_to_15min(dt):
    """将时间向上取整到最近的15分钟"""
    minutes = dt.minute
    rounded_minutes = ((minutes + 14) // 15) * 15
    return dt + timedelta(minutes=rounded_minutes - minutes)

def apply_excel_styles(file_path):
    """应用Excel样式：自适应列宽、表头黄色背景和加粗、苹果平方字体、黑色边框"""
    print(f"{Fore.CYAN}正在应用Excel样式，这可能需要一些时间...{Fore.RESET}")
    
    # 加载工作簿
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    ws = wb.active
    
    # 定义样式
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(name="PingFang SC", bold=True, size=11)
    cell_font = Font(name="PingFang SC", size=11)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 获取数据范围
    data_rows = min(ws.max_row, 10000)  # 限制最大处理行数为10000行
    data_cols = ws.max_column
    
    # 应用表头样式（只处理第一行）
    print(f"{Fore.CYAN}应用表头样式...{Fore.RESET}")
    for col in range(1, data_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 批量设置字体和边框（跳过表头）
    print(f"{Fore.CYAN}应用数据单元格样式...{Fore.RESET}")
    for row in tqdm(range(2, data_rows + 1), desc="应用样式"):
        for col in range(1, data_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 自适应列宽
    print(f"{Fore.CYAN}调整列宽...{Fore.RESET}")
    for col in range(1, data_cols + 1):
        max_length = 0
        column = get_column_letter(col)
        
        # 只检查前100行来确定列宽
        for row in range(1, min(100, data_rows + 1)):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                try:
                    if len(str(cell_value)) > max_length:
                        max_length = len(str(cell_value))
                except:
                    pass
        
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = min(adjusted_width, 50)  # 限制最大宽度为50
    
    # 保存修改后的工作簿
    print(f"{Fore.CYAN}保存样式化的Excel文件...{Fore.RESET}")
    wb.save(file_path)
    print(f"{Fore.GREEN}已应用Excel样式美化到：{file_path}{Fore.RESET}")

def process_post_list():
    print(f"{Fore.CYAN}开始处理发帖列表数据...{Fore.RESET}")
    
    # 获取所有post_list文件
    files = glob.glob(str(RAW_DIR / "bbs_post_list_*.xlsx"))
    if not files:
        print(f"{Fore.RED}错误：未找到发帖列表数据文件！{Fore.RESET}")
        return
    
    all_data = []
    
    # 处理每个文件
    for file in tqdm(files, desc="处理文件"):
        # 读取所有工作表
        excel_file = pd.ExcelFile(file)
        source_file = os.path.basename(file)
        
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            # 添加source_file和sheet_name字段
            df['source_file'] = source_file
            df['sheet_name'] = sheet_name
            
            # 计算scraping_time_R
            min_scraping_time = pd.to_datetime(df['scraping_time']).min()
            scraping_time_r = round_time_to_15min(min_scraping_time)
            df['scraping_time_R'] = scraping_time_r
            
            all_data.append(df)
    
    # 合并所有数据
    df_combined = pd.concat(all_data, ignore_index=True)
    
    # 处理post_time
    df_combined = df_combined.sort_values('scraping_time')
    
    # 处理scraping_time
    df_combined['scraping_time'] = pd.to_datetime(df_combined['scraping_time'])
    
    # 处理list_time，先尝试直接转换
    def convert_datetime(x):
        try:
            # 如果是完整的日期时间格式
            return pd.to_datetime(x)
        except:
            try:
                # 如果只有日期，添加时间00:00:00
                if isinstance(x, str) and len(x.strip()) <= 10:
                    return pd.to_datetime(x.strip() + ' 00:00:00')
                return pd.to_datetime(x)
            except:
                return pd.NaT
    
    df_combined['list_time'] = df_combined['list_time'].apply(convert_datetime)
    
    # 对于只有日期的情况，补充时间为00:00:00
    def complete_time(x):
        if pd.notna(x):
            try:
                if pd.to_datetime(x).strftime('%H:%M:%S') == '00:00:00':
                    return x
                return x
            except:
                return pd.to_datetime(str(x).split()[0] + ' 00:00:00')
        return x
    
    df_combined['post_time'] = df_combined.groupby('url')['list_time'].transform('first').apply(complete_time)
    
    # 只保留每个url最早的记录
    df_result = df_combined.sort_values('scraping_time').groupby('url').first().reset_index()
    
    # 设置输出列顺序
    columns_order = [
        'url', 'title', 'scraping_time_R', 'post_time', 'list_time', 'page', 'num',
        'author', 'author_link', 'read_count', 'reply_count', 'scraping_time',
        'list_time', 'source_file', 'sheet_name'
    ]
    df_result = df_result[columns_order]
    
    # 校验数据
    if df_result.isnull().any().any():
        print(f"{Fore.RED}警告：输出数据中存在空值！{Fore.RESET}")
        return
    
    # 确保时间格式正确
    time_columns = ['scraping_time_R', 'post_time']
    for col in time_columns:
        if not all(df_result[col].astype(str).str.contains(' ')):
            print(f"{Fore.RED}警告：{col}列存在不完整的时间格式！{Fore.RESET}")
            return
    
    # 创建输出目录
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    
    # 保存结果
    output_file = PROCESSED_DIR / 'post.xlsx'
    df_result.to_excel(output_file, index=False)
    
    # 应用Excel样式美化
    apply_excel_styles(output_file)
    
    print(f"{Fore.GREEN}处理完成！结果已保存至：{output_file}{Fore.RESET}")

if __name__ == "__main__":
    process_post_list() 