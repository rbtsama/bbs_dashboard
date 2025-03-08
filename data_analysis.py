import os
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from tqdm import tqdm
from colorama import init, Fore
import calendar
import traceback
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# 初始化colorama
init()

def get_week_range(date):
    """获取日期所在的自然周范围（周一至周日）"""
    # 获取周一日期
    start_of_week = date - timedelta(days=date.weekday())
    # 获取周日日期
    end_of_week = start_of_week + timedelta(days=6)
    return f"{start_of_week.strftime('%Y-%m-%d')}~{end_of_week.strftime('%Y-%m-%d')}"

def get_month_year(date):
    """获取年月格式 YYYY-MM"""
    return date.strftime('%Y-%m')

def hour_format(hour):
    """格式化小时为中文表示"""
    return f"{hour}点"

def ensure_all_hours(df, date_col, hour_col, value_cols, group_col=None, fill_value=0):
    """确保所有时间段都有记录，即使计数为0"""
    # 创建有序的小时列表
    ordered_hours = [hour_format(h) for h in range(24)]
    
    # 获取所有唯一日期/分组
    if group_col:
        all_groups = df[group_col].unique()
        all_combinations = []
        for group in all_groups:
            for hour in ordered_hours:
                all_combinations.append({
                    group_col: group,
                    hour_col: hour
                })
    else:
        all_dates = df[date_col].unique()
        all_combinations = []
        for date in all_dates:
            for hour in ordered_hours:
                all_combinations.append({
                    date_col: date,
                    hour_col: hour
                })
    
    # 创建完整的DataFrame
    complete_df = pd.DataFrame(all_combinations)
    
    # 合并原始数据
    if group_col:
        merged_df = pd.merge(complete_df, df, on=[group_col, hour_col], how='left')
    else:
        merged_df = pd.merge(complete_df, df, on=[date_col, hour_col], how='left')
    
    # 填充缺失值
    for col in value_cols:
        if col in merged_df.columns:
            merged_df[col] = merged_df[col].fillna(fill_value)
    
    return merged_df

def analyze_post_time_distribution(post_df):
    """分析一：发帖时间分布分析"""
    print(f"{Fore.CYAN}开始分析发帖时间分布...{Fore.RESET}")
    
    # 确保post_time是datetime类型
    post_df['post_time'] = pd.to_datetime(post_df['post_time'])
    
    # 提取日期和小时
    post_df['date'] = post_df['post_time'].dt.date
    post_df['hour'] = post_df['post_time'].dt.hour
    post_df['hour_str'] = post_df['hour'].apply(hour_format)
    
    # 添加周和月信息
    post_df['week_range'] = post_df['post_time'].apply(lambda x: get_week_range(x))
    post_df['year_month'] = post_df['post_time'].apply(lambda x: get_month_year(x))
    
    # 创建有序的小时列表
    ordered_hours = [hour_format(h) for h in range(24)]
    
    # 1. 日发帖量统计表
    daily_post_counts = post_df.groupby(['date', 'hour_str']).size().reset_index(name='count')
    
    # 确保所有小时段都有记录
    daily_post_counts = ensure_all_hours(
        daily_post_counts, 'date', 'hour_str', ['count']
    )
    
    # 2. 周发帖量统计表
    weekly_post_counts = post_df.groupby(['week_range', 'hour_str']).size().reset_index(name='count')
    
    # 确保所有周的所有小时段都有记录
    weekly_post_counts = ensure_all_hours(
        weekly_post_counts, None, 'hour_str', ['count'], 
        group_col='week_range', fill_value=0
    )
    
    # 3. 月发帖量统计表
    monthly_post_counts = post_df.groupby(['year_month', 'hour_str']).size().reset_index(name='count')
    
    # 确保所有月份的所有小时段都有记录
    monthly_post_counts = ensure_all_hours(
        monthly_post_counts, None, 'hour_str', ['count'],
        group_col='year_month', fill_value=0
    )
    
    # 4. 时段发帖汇总表
    hourly_post_counts = post_df.groupby('hour_str').size().reset_index(name='count')
    
    # 确保所有小时段都有记录
    all_hours = pd.DataFrame({'hour_str': ordered_hours})
    hourly_post_counts = pd.merge(all_hours, hourly_post_counts, on='hour_str', how='left').fillna(0)
    hourly_post_counts['count'] = hourly_post_counts['count'].astype(int)
    
    return {
        'daily_post_counts': daily_post_counts,
        'weekly_post_counts': weekly_post_counts,
        'monthly_post_counts': monthly_post_counts,
        'hourly_post_counts': hourly_post_counts
    }

def analyze_update_behavior(update_df):
    """分析二：帖子更新行为分析"""
    print(f"{Fore.CYAN}开始分析帖子更新行为...{Fore.RESET}")
    
    # 确保list_time_R是datetime类型
    update_df['list_time_R'] = pd.to_datetime(update_df['list_time_R'])
    
    # 提取日期和小时
    update_df['date'] = update_df['list_time_R'].dt.date
    update_df['hour'] = update_df['list_time_R'].dt.hour
    update_df['hour_str'] = update_df['hour'].apply(hour_format)
    
    # 添加周和月信息
    update_df['week_range'] = update_df['list_time_R'].apply(lambda x: get_week_range(x))
    update_df['year_month'] = update_df['list_time_R'].apply(lambda x: get_month_year(x))
    
    # 过滤有更新原因的记录
    update_df_filtered = update_df[update_df['update_reason'].notna() & (update_df['update_reason'] != '')]
    
    # 创建有序的小时列表
    ordered_hours = [hour_format(h) for h in range(24)]
    
    # 1. 日更新量统计表
    update_reasons = ['重发', '回帖', '删回帖']
    daily_update_counts = pd.DataFrame()
    
    for reason in update_reasons:
        reason_df = update_df_filtered[update_df_filtered['update_reason'] == reason]
        reason_counts = reason_df.groupby(['date', 'hour_str']).size().reset_index(name=reason)
        
        if daily_update_counts.empty:
            daily_update_counts = reason_counts
        else:
            daily_update_counts = pd.merge(
                daily_update_counts, reason_counts, 
                on=['date', 'hour_str'], how='outer'
            )
    
    # 确保所有小时段都有记录
    daily_update_counts = ensure_all_hours(
        daily_update_counts, 'date', 'hour_str', update_reasons
    )
    
    # 2. 周更新量统计表
    weekly_update_counts = pd.DataFrame()
    
    for reason in update_reasons:
        reason_df = update_df_filtered[update_df_filtered['update_reason'] == reason]
        reason_counts = reason_df.groupby(['week_range', 'hour_str']).size().reset_index(name=reason)
        
        if weekly_update_counts.empty:
            weekly_update_counts = reason_counts
        else:
            weekly_update_counts = pd.merge(
                weekly_update_counts, reason_counts, 
                on=['week_range', 'hour_str'], how='outer'
            )
    
    # 确保所有时间段都有记录
    weekly_update_counts = ensure_all_hours(
        weekly_update_counts, None, 'hour_str', update_reasons, 
        group_col='week_range', fill_value=0
    )
    
    # 3. 月更新量统计表
    monthly_update_counts = pd.DataFrame()
    
    for reason in update_reasons:
        reason_df = update_df_filtered[update_df_filtered['update_reason'] == reason]
        reason_counts = reason_df.groupby(['year_month', 'hour_str']).size().reset_index(name=reason)
        
        if monthly_update_counts.empty:
            monthly_update_counts = reason_counts
        else:
            monthly_update_counts = pd.merge(
                monthly_update_counts, reason_counts, 
                on=['year_month', 'hour_str'], how='outer'
            )
    
    # 确保所有时间段都有记录
    monthly_update_counts = ensure_all_hours(
        monthly_update_counts, None, 'hour_str', update_reasons,
        group_col='year_month', fill_value=0
    )
    
    # 4. 时段更新汇总表
    hourly_update_counts = pd.DataFrame()
    
    for reason in update_reasons:
        reason_df = update_df_filtered[update_df_filtered['update_reason'] == reason]
        reason_counts = reason_df.groupby('hour_str').size().reset_index(name=reason)
        
        if hourly_update_counts.empty:
            hourly_update_counts = reason_counts
        else:
            hourly_update_counts = pd.merge(
                hourly_update_counts, reason_counts, 
                on='hour_str', how='outer'
            )
    
    # 确保所有小时段都有记录
    all_hours = pd.DataFrame({'hour_str': ordered_hours})
    hourly_update_counts = pd.merge(all_hours, hourly_update_counts, on='hour_str', how='left').fillna(0)
    
    # 转换为整数
    for reason in update_reasons:
        hourly_update_counts[reason] = hourly_update_counts[reason].astype(int)
    
    return {
        'daily_update_counts': daily_update_counts,
        'weekly_update_counts': weekly_update_counts,
        'monthly_update_counts': monthly_update_counts,
        'hourly_update_counts': hourly_update_counts
    }

def analyze_read_count_growth(update_df):
    """分析三：阅读量增长分析"""
    print(f"{Fore.CYAN}开始分析阅读量增长...{Fore.RESET}")
    
    # 确保scraping_time_R是datetime类型
    update_df['scraping_time_R'] = pd.to_datetime(update_df['scraping_time_R'])
    
    # 按日期、小时和URL分组，获取每组中最接近15分钟的记录
    def get_nearest_snapshot(group):
        if len(group) == 0:
            return None
        # 计算每个时间点与15分钟的差值
        group['time_diff'] = abs(group['scraping_time_R'].dt.minute - 15)
        # 获取差值最小的记录
        return group.loc[group['time_diff'].idxmin()]
    
    # 按日期、小时和URL分组
    snapshot_df = update_df.groupby([
        update_df['scraping_time_R'].dt.date,
        update_df['scraping_time_R'].dt.hour,
        'url'
    ]).apply(get_nearest_snapshot).reset_index(drop=True)
    
    if len(snapshot_df) == 0:
        # 如果没有数据，创建空的结果DataFrame
        empty_dates = update_df['scraping_time_R'].dt.date.unique()
        empty_hours = [hour_format(h) for h in range(24)]
        empty_combinations = []
        for date in empty_dates:
            for hour in empty_hours:
                empty_combinations.append({
                    'date': date,
                    'hour_str': hour,
                    'read_gap': 0
                })
        return {
            'daily_read_gap': pd.DataFrame(empty_combinations),
            'monthly_avg_read_gap': pd.DataFrame(columns=['year_month', 'hour_str', 'avg_read_gap'])
        }
    
    # 提取日期和小时
    snapshot_df['date'] = snapshot_df['scraping_time_R'].dt.date
    snapshot_df['hour'] = snapshot_df['scraping_time_R'].dt.hour
    snapshot_df['hour_str'] = snapshot_df['hour'].apply(hour_format)
    snapshot_df['year_month'] = snapshot_df['scraping_time_R'].apply(lambda x: get_month_year(x))
    
    # 按URL和时间排序
    snapshot_df = snapshot_df.sort_values(['url', 'scraping_time_R'])
    
    # 计算相邻时段相同URL帖子的阅读量差值
    snapshot_df['prev_read_count'] = snapshot_df.groupby('url')['read_count'].shift(1)
    snapshot_df['read_gap'] = snapshot_df['read_count'] - snapshot_df['prev_read_count']
    
    # 过滤掉无效的差值（第一条记录或负值）
    snapshot_df = snapshot_df[snapshot_df['read_gap'].notna() & (snapshot_df['read_gap'] >= 0)]
    
    # 1. 日阅读量增量表
    daily_read_gap = snapshot_df.groupby(['date', 'hour_str'])['read_gap'].sum().reset_index()
    daily_read_gap['read_gap'] = daily_read_gap['read_gap'].astype(int)
    
    # 确保所有小时段都有记录
    daily_read_gap = ensure_all_hours(
        daily_read_gap, 'date', 'hour_str', ['read_gap']
    )
    daily_read_gap['read_gap'] = daily_read_gap['read_gap'].astype(int)
    
    # 2. 月阅读量平均增量表 - 修正计算方法
    # 首先按月份、小时和日期分组，计算每天每小时的总增量
    daily_by_month = snapshot_df.groupby(['year_month', 'hour_str', 'date'])['read_gap'].sum().reset_index()
    
    # 然后按月份和小时分组，计算每个月每个小时的平均值（基于有数据的天数）
    monthly_avg_read_gap = daily_by_month.groupby(['year_month', 'hour_str'])['read_gap'].mean().reset_index()
    monthly_avg_read_gap['avg_read_gap'] = monthly_avg_read_gap['read_gap'].round().astype(int)
    monthly_avg_read_gap = monthly_avg_read_gap.drop(columns=['read_gap'])
    
    # 确保所有月份的所有小时段都有记录
    monthly_avg_read_gap = ensure_all_hours(
        monthly_avg_read_gap, None, 'hour_str', ['avg_read_gap'],
        group_col='year_month', fill_value=0
    )
    
    return {
        'daily_read_gap': daily_read_gap,
        'monthly_avg_read_gap': monthly_avg_read_gap
    }

def analyze_post_activity(update_df, post_df):
    """分析四：帖子活跃度排行榜"""
    print(f"{Fore.CYAN}开始分析帖子活跃度...{Fore.RESET}")
    
    # 确保post_time是datetime类型
    post_df['post_time'] = pd.to_datetime(post_df['post_time'])
    update_df['list_time_R'] = pd.to_datetime(update_df['list_time_R'])
    
    # 按URL汇总各种更新行为次数
    update_reasons = ['重发', '回帖', '删回帖']
    post_activity = pd.DataFrame()
    
    # 首先获取所有出现在update_df中的URL
    all_urls = update_df['url'].unique()
    post_activity['url'] = all_urls
    
    for reason in update_reasons:
        reason_df = update_df[update_df['update_reason'] == reason]
        reason_counts = reason_df.groupby('url').size().reset_index(name=reason)
        post_activity = pd.merge(post_activity, reason_counts, on='url', how='left')
    
    # 合并帖子基本信息，使用left join保留所有URL
    post_info = post_df[['url', 'post_time', 'title', 'author']].drop_duplicates(subset=['url'])
    post_activity = pd.merge(post_activity, post_info, on='url', how='left')
    
    # 获取每个URL最近一次的list_time_R
    latest_list_times = update_df.groupby('url')['list_time_R'].max().reset_index()
    latest_list_times.rename(columns={'list_time_R': 'latest_list_time'}, inplace=True)
    post_activity = pd.merge(post_activity, latest_list_times, on='url', how='left')
    
    # 对于缺失的post_time，从update_df中获取最早的list_time_R作为近似
    for idx, row in post_activity.iterrows():
        if pd.isna(row['post_time']):
            url = row['url']
            # 获取该URL的所有更新记录
            url_updates = update_df[update_df['url'] == url]
            if len(url_updates) > 0:
                # 使用最早的list_time_R作为近似的post_time
                earliest_time = url_updates['list_time_R'].min()
                post_activity.at[idx, 'post_time'] = earliest_time
    
    # 对于缺失的title和author，从update_df中获取
    update_info = update_df[['url', 'title', 'author']].drop_duplicates(subset=['url'])
    
    # 填充缺失的title和author
    post_activity['title'] = post_activity.apply(
        lambda x: x['title'] if pd.notna(x['title']) 
        else update_info[update_info['url'] == x['url']]['title'].iloc[0] 
        if len(update_info[update_info['url'] == x['url']]) > 0 
        else '未知标题',
        axis=1
    )
    
    post_activity['author'] = post_activity.apply(
        lambda x: x['author'] if pd.notna(x['author']) 
        else update_info[update_info['url'] == x['url']]['author'].iloc[0]
        if len(update_info[update_info['url'] == x['url']]) > 0 
        else '未知作者',
        axis=1
    )
    
    # 计算发帖天数，处理可能的NaT值
    latest_date = post_df['post_time'].max()
    if pd.isna(latest_date):
        latest_date = update_df['list_time_R'].max()
    latest_date = latest_date.date()
    
    def calculate_days_old(x):
        if pd.isna(x) or pd.isnull(x):
            return 0
        try:
            return (latest_date - x.date()).days
        except:
            return 0
    
    post_activity['daysold'] = post_activity['post_time'].apply(calculate_days_old)
    
    # 填充缺失值
    post_activity = post_activity.fillna({reason: 0 for reason in update_reasons})
    
    # 按重发次数降序排列
    post_activity = post_activity.sort_values('重发', ascending=False)
    
    # 选择需要的列
    post_activity = post_activity[['url', 'post_time', 'latest_list_time', 'title', 'daysold', 'author', '重发', '回帖', '删回帖']]
    
    return post_activity

def analyze_user_activity(update_df):
    """分析五：用户活跃度排行榜"""
    print(f"{Fore.CYAN}开始分析用户活跃度...{Fore.RESET}")
    
    # 按作者汇总各种更新行为次数
    update_reasons = ['重发', '回帖', '删回帖']
    user_activity = pd.DataFrame()
    
    for reason in update_reasons:
        reason_df = update_df[update_df['update_reason'] == reason]
        reason_counts = reason_df.groupby('author').size().reset_index(name=reason)
        
        if user_activity.empty:
            user_activity = reason_counts
        else:
            user_activity = pd.merge(
                user_activity, reason_counts, 
                on='author', how='outer'
            )
    
    # 填充缺失值
    user_activity = user_activity.fillna(0)
    
    # 按重发次数降序排列
    user_activity = user_activity.sort_values('重发', ascending=False)
    
    # 转换为整数
    for reason in update_reasons:
        user_activity[reason] = user_activity[reason].astype(int)
    
    return user_activity

def apply_excel_styles(writer, sheet_name):
    """应用Excel样式：自适应列宽、表头黄色背景和加粗、苹果平方字体、黑色边框"""
    print(f"{Fore.CYAN}正在为'{sheet_name}'应用Excel样式...{Fore.RESET}")
    
    # 获取工作表
    worksheet = writer.sheets[sheet_name]
    
    # 定义样式
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(name="PingFang SC", bold=True, size=11)
    cell_font = Font(name="PingFang SC", size=11)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 获取数据范围
    data_rows = min(worksheet.max_row, 10000)  # 限制最大处理行数为10000行
    data_cols = worksheet.max_column
    
    # 应用表头样式（只处理第一行）
    for col in range(1, data_cols + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 批量设置字体和边框（跳过表头）
    for row in range(2, data_rows + 1):
        for col in range(1, data_cols + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 自适应列宽
    for col in range(1, data_cols + 1):
        max_length = 0
        column = get_column_letter(col)
        
        # 只检查前100行来确定列宽
        for row in range(1, min(100, data_rows + 1)):
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value:
                try:
                    if len(str(cell_value)) > max_length:
                        max_length = len(str(cell_value))
                except:
                    pass
        
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = min(adjusted_width, 50)  # 限制最大宽度为50

def run_analysis():
    """运行所有分析"""
    # 设置进度条
    total_steps = 5
    progress_bar = tqdm(total=total_steps, desc="总体进度")
    
    # 读取预处理后的数据
    try:
        post_df = pd.read_excel('output/post_pre.xlsx')
        update_df = pd.read_excel('output/update_pre.xlsx')
    except Exception as e:
        print(f"{Fore.RED}错误：无法读取数据文件 - {e}{Fore.RESET}")
        return
    
    # 分析一：发帖时间分布分析
    post_time_results = analyze_post_time_distribution(post_df)
    progress_bar.update(1)
    
    # 分析二：帖子更新行为分析
    update_behavior_results = analyze_update_behavior(update_df)
    progress_bar.update(1)
    
    # 分析三：阅读量增长分析
    read_growth_results = analyze_read_count_growth(update_df)
    progress_bar.update(1)
    
    # 分析四：帖子活跃度排行榜
    post_activity = analyze_post_activity(update_df, post_df)
    progress_bar.update(1)
    
    # 分析五：用户活跃度排行榜
    user_activity = analyze_user_activity(update_df)
    progress_bar.update(1)
    
    # 保存结果到Excel
    print(f"{Fore.GREEN}保存分析结果...{Fore.RESET}")
    try:
        # 确保output目录存在
        os.makedirs('output', exist_ok=True)
        
        with pd.ExcelWriter('output/data_analysis.xlsx', engine='openpyxl') as writer:
            # 创建有序的小时列表
            ordered_hours = [hour_format(h) for h in range(24)]
            
            # 分析一：发帖时间分布分析
            # 日发帖量统计表 - 透视表格式
            daily_pivot = post_time_results['daily_post_counts'].pivot(
                index='date', columns='hour_str', values='count'
            ).fillna(0).astype(int)
            # 重新排序列
            daily_pivot = daily_pivot.reindex(columns=ordered_hours)
            # 添加总计列
            daily_pivot['总计'] = daily_pivot.sum(axis=1)
            daily_pivot.to_excel(writer, sheet_name='日发帖量统计表')
            apply_excel_styles(writer, '日发帖量统计表')
            
            # 周发帖量统计表 - 透视表格式
            weekly_pivot = post_time_results['weekly_post_counts'].pivot(
                index='week_range', columns='hour_str', values='count'
            ).fillna(0).astype(int)
            weekly_pivot = weekly_pivot.reindex(columns=ordered_hours)
            weekly_pivot['总计'] = weekly_pivot.sum(axis=1)
            weekly_pivot.to_excel(writer, sheet_name='周发帖量统计表')
            apply_excel_styles(writer, '周发帖量统计表')
            
            # 月发帖量统计表 - 透视表格式
            monthly_pivot = post_time_results['monthly_post_counts'].pivot(
                index='year_month', columns='hour_str', values='count'
            ).fillna(0).astype(int)
            monthly_pivot = monthly_pivot.reindex(columns=ordered_hours)
            monthly_pivot['总计'] = monthly_pivot.sum(axis=1)
            monthly_pivot.to_excel(writer, sheet_name='月发帖量统计表')
            apply_excel_styles(writer, '月发帖量统计表')
            
            # 时段发帖汇总表
            hourly_post = post_time_results['hourly_post_counts'].copy()
            # 添加总计行
            total_row = pd.DataFrame({
                'hour_str': ['总计'],
                'count': [hourly_post['count'].sum()]
            })
            hourly_post = pd.concat([hourly_post, total_row], ignore_index=True)
            hourly_post.to_excel(writer, sheet_name='时段发帖汇总表', index=False)
            apply_excel_styles(writer, '时段发帖汇总表')
            
            # 分析二：帖子更新行为分析
            # 日更新量统计表 - 透视表格式
            daily_update_pivot = update_behavior_results['daily_update_counts'].pivot(
                index='date', columns='hour_str', 
                values=['重发', '回帖', '删回帖']
            ).fillna(0).astype(int)
            
            # 重新排序列，确保按照0点、1点、2点...23点排列
            # 创建多级列索引的排序顺序
            update_cols = ['重发', '回帖', '删回帖']
            ordered_cols = pd.MultiIndex.from_product([update_cols, ordered_hours])
            
            # 重新排序列
            daily_update_pivot = daily_update_pivot.reindex(columns=ordered_cols)
            
            # 添加总计列
            for col in update_cols:
                daily_update_pivot[(col, '总计')] = daily_update_pivot[col].sum(axis=1)
            
            daily_update_pivot.to_excel(writer, sheet_name='日更新量统计表')
            apply_excel_styles(writer, '日更新量统计表')
            
            # 周更新量统计表 - 透视表格式
            weekly_update_pivot = update_behavior_results['weekly_update_counts'].pivot(
                index='week_range', columns='hour_str', 
                values=['重发', '回帖', '删回帖']
            ).fillna(0).astype(int)
            
            # 重新排序列
            weekly_update_pivot = weekly_update_pivot.reindex(columns=ordered_cols)
            
            # 计算总计
            for col in update_cols:
                weekly_update_pivot[(col, '总计')] = weekly_update_pivot[col].sum(axis=1)
            
            weekly_update_pivot.to_excel(writer, sheet_name='周更新量统计表')
            apply_excel_styles(writer, '周更新量统计表')
            
            # 月更新量统计表 - 透视表格式
            monthly_update_pivot = update_behavior_results['monthly_update_counts'].pivot(
                index='year_month', columns='hour_str', 
                values=['重发', '回帖', '删回帖']
            ).fillna(0).astype(int)
            
            # 重新排序列
            monthly_update_pivot = monthly_update_pivot.reindex(columns=ordered_cols)
            
            # 计算总计
            for col in update_cols:
                monthly_update_pivot[(col, '总计')] = monthly_update_pivot[col].sum(axis=1)
            
            monthly_update_pivot.to_excel(writer, sheet_name='月更新量统计表')
            apply_excel_styles(writer, '月更新量统计表')
            
            # 时段更新汇总表
            # 确保按照0点、1点、2点...23点排列
            hourly_update = update_behavior_results['hourly_update_counts'].copy()
            # 创建小时顺序映射
            hour_order = {hour: i for i, hour in enumerate(ordered_hours)}
            # 添加排序键
            hourly_update['hour_order'] = hourly_update['hour_str'].map(hour_order)
            # 排序并删除排序键
            hourly_update = hourly_update.sort_values('hour_order').drop('hour_order', axis=1)
            
            # 添加总计行
            total_row = pd.DataFrame({
                'hour_str': ['总计'],
                '重发': [hourly_update['重发'].sum()],
                '回帖': [hourly_update['回帖'].sum()],
                '删回帖': [hourly_update['删回帖'].sum()]
            })
            hourly_update = pd.concat([hourly_update, total_row], ignore_index=True)
            
            hourly_update.to_excel(writer, sheet_name='时段更新汇总表', index=False)
            apply_excel_styles(writer, '时段更新汇总表')
            
            # 分析三：阅读量增长分析
            # 日阅读量增量表 - 透视表格式
            daily_read_pivot = read_growth_results['daily_read_gap'].pivot(
                index='date', columns='hour_str', values='read_gap'
            ).fillna(0).astype(int)
            daily_read_pivot = daily_read_pivot.reindex(columns=ordered_hours)
            # 添加总计列
            daily_read_pivot['总计'] = daily_read_pivot.sum(axis=1)
            daily_read_pivot.to_excel(writer, sheet_name='日阅读量增量表')
            apply_excel_styles(writer, '日阅读量增量表')
            
            # 月阅读量平均增量表 - 透视表格式
            monthly_read_pivot = read_growth_results['monthly_avg_read_gap'].pivot(
                index='year_month', columns='hour_str', values='avg_read_gap'
            ).fillna(0).astype(int)
            monthly_read_pivot = monthly_read_pivot.reindex(columns=ordered_hours)
            # 添加总计列 - 对于平均值，计算所有小时的平均值
            monthly_read_pivot['总计'] = monthly_read_pivot.mean(axis=1).round().astype(int)
            monthly_read_pivot.to_excel(writer, sheet_name='月阅读量平均增量表')
            apply_excel_styles(writer, '月阅读量平均增量表')
            
            # 分析四：帖子活跃度排行榜
            post_activity.to_excel(writer, sheet_name='帖子活跃度排行榜', index=False)
            apply_excel_styles(writer, '帖子活跃度排行榜')
            
            # 分析五：用户活跃度排行榜
            user_activity.to_excel(writer, sheet_name='用户活跃度排行榜', index=False)
            apply_excel_styles(writer, '用户活跃度排行榜')
        
        print(f"{Fore.GREEN}分析完成！结果已保存至：output/data_analysis.xlsx{Fore.RESET}")
    except Exception as e:
        print(f"{Fore.RED}错误：保存结果失败 - {e}{Fore.RESET}")
        print(f"错误详情: {traceback.format_exc()}")

if __name__ == "__main__":
    # 设置控制台颜色
    init(autoreset=True)
    
    # 运行分析
    run_analysis() 